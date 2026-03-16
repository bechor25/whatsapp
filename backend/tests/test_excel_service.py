"""
Unit tests for ExcelService — phone normalisation, duplicate detection, etc.
No network or WhatsApp required.
"""
import io
import pytest
import openpyxl
from services.excel_service import ExcelService


@pytest.fixture
def svc():
    return ExcelService()


def _make_xlsx(rows: list) -> io.BytesIO:
    """Create an in-memory xlsx with the given (name, phone) rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Basic parsing ─────────────────────────────────────────────────────────────
# parse_excel() returns (contacts: list, errors: list)

class TestParseBasics:
    def test_three_contacts(self, svc, tmp_path):
        rows = [("ישראל ישראלי", "0501234567"),
                ("שרה כהן",      "0521234567"),
                ("דוד לוי",      "0541234567")]
        buf = _make_xlsx(rows)
        path = tmp_path / "c.xlsx"
        path.write_bytes(buf.read())

        contacts, errors = svc.parse_excel(str(path))
        assert len(contacts) == 3

    def test_names_preserved(self, svc, tmp_path):
        rows = [("אבי", "0501111111")]
        buf = _make_xlsx(rows)
        path = tmp_path / "c.xlsx"
        path.write_bytes(buf.read())

        contacts, _ = svc.parse_excel(str(path))
        assert contacts[0]["name"] == "אבי"

    def test_empty_rows_skipped(self, svc, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["שם א", "0501111111"])
        ws.append(["", ""])           # empty row
        ws.append(["שם ב", "0502222222"])
        path = tmp_path / "gaps.xlsx"
        wb.save(str(path))

        contacts, _ = svc.parse_excel(str(path))
        assert len(contacts) == 2

    def test_file_not_found_raises(self, svc):
        with pytest.raises(Exception):
            svc.parse_excel("/nosuchfile.xlsx")


# ── Phone normalisation ───────────────────────────────────────────────────────
# _normalize_phone(phone, row_num) → (normalized_str | None, error_str | None)

class TestPhoneNormalisation:
    def _norm(self, svc, phone: str):
        """Return only the normalised string (or None)."""
        normalised, _ = svc._normalize_phone(phone, 1)
        return normalised

    def test_local_050(self, svc):
        assert self._norm(svc, "0501234567") == "972501234567"

    def test_local_052(self, svc):
        assert self._norm(svc, "0521234567") == "972521234567"

    def test_plus_972(self, svc):
        assert self._norm(svc, "+972501234567") == "972501234567"

    def test_972_no_plus(self, svc):
        assert self._norm(svc, "972501234567") == "972501234567"

    def test_with_dashes(self, svc):
        assert self._norm(svc, "050-123-4567") == "972501234567"

    def test_with_spaces(self, svc):
        assert self._norm(svc, "050 123 4567") == "972501234567"

    def test_landline_not_mobile_does_not_crash(self, svc):
        result = self._norm(svc, "021234567")
        assert result is None or isinstance(result, str)

    def test_too_short_returns_none(self, svc):
        assert self._norm(svc, "050") is None

    def test_empty_returns_none(self, svc):
        result = self._norm(svc, "")
        assert result is None

    def test_error_message_on_invalid(self, svc):
        _, err = svc._normalize_phone("999", 5)
        assert err is not None
        assert "5" in err  # should mention the row number


# ── Duplicate detection ───────────────────────────────────────────────────────

class TestDuplicates:
    def test_duplicate_phone_removed(self, svc, tmp_path):
        rows = [("אחד", "0501234567"),
                ("שניים", "0501234567")]   # same phone
        buf = _make_xlsx(rows)
        path = tmp_path / "d.xlsx"
        path.write_bytes(buf.read())

        contacts, errors = svc.parse_excel(str(path))
        assert len(contacts) == 1
        assert len(errors) >= 1

    def test_different_names_different_phones_both_kept(self, svc, tmp_path):
        rows = [("אבי", "0501234567"),
                ("שרה", "0521234567")]
        buf = _make_xlsx(rows)
        path = tmp_path / "two.xlsx"
        path.write_bytes(buf.read())

        contacts, errors = svc.parse_excel(str(path))
        assert len(contacts) == 2
        assert len(errors) == 0


# ── Validation errors ─────────────────────────────────────────────────────────

class TestValidationErrors:
    def test_missing_name_reported_in_errors(self, svc, tmp_path):
        rows = [("", "0501234567")]
        buf = _make_xlsx(rows)
        path = tmp_path / "noname.xlsx"
        path.write_bytes(buf.read())

        contacts, errors = svc.parse_excel(str(path))
        assert len(contacts) == 0
        assert len(errors) >= 1

    def test_missing_phone_reported_in_errors(self, svc, tmp_path):
        # Write the phone cell explicitly so pandas sees 2 columns
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="פלוני אלמוני")
        ws.cell(row=1, column=2, value=" ")   # space → treated as empty after strip
        path = tmp_path / "nophone.xlsx"
        wb.save(str(path))

        contacts, errors = svc.parse_excel(str(path))
        assert len(contacts) == 0
        assert len(errors) >= 1

    def test_invalid_phone_reported_in_errors(self, svc, tmp_path):
        # Use a non-keyword name so it isn't mistaken for a header row
        rows = [("אבי שמיר", "999")]
        buf = _make_xlsx(rows)
        path = tmp_path / "bad.xlsx"
        path.write_bytes(buf.read())

        contacts, errors = svc.parse_excel(str(path))
        assert len(contacts) == 0
        assert len(errors) >= 1
